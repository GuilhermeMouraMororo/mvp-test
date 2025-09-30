import re
import unicodedata
from copy import deepcopy
import time
import threading
import os
from flask import Flask, render_template, request, jsonify, send_file, session
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'

# --------- Excel / persistence setup ----------
EXCEL_PATH = "Pedidos.xlsx"
excel_lock = threading.Lock()

# Load or initialize Excel
try:
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active
    ws.append(["Produto", "Quantidade"])  # header
    wb.save(EXCEL_PATH)

# ---------- Your Core Order Processing Functions ----------

def normalize(text):
    text = text.lower()
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')
    return text.strip() 

def levenshtein_distance(a, b):
    m, n = len(a), len(b)
    if m == 0: return n
    if n == 0: return m
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(m + 1):
        dp[i][0] = i
    for j in range(n + 1):
        dp[0][j] = j    
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            dp[i][j] = min(
                dp[i - 1][j] + 1,
                dp[i][j - 1] + 1,
                dp[i - 1][j - 1] + cost
            )
    return dp[m][n]

def similarity_percentage(a, b):
    a, b = normalize(a), normalize(b)
    distance = levenshtein_distance(a, b)
    max_len = max(len(a), len(b))
    if max_len == 0:
        return 100.0
    return (1 - distance / max_len) * 100

units = {
    "zero":0, "um":1, "uma":1, "dois":2, "duas":2, "dos":2, "tres":3, "tres":3, "treis": 3,
    "quatro":4, "quarto":4, "cinco":5, "cnico": 5, "seis":6, "ses":6, "sete":7, "oito":8, "nove":9, "nov": 9
}
teens = {
    "dez":10, "onze":11, "doze":12, "treze":13, "quatorze":14, "catorze":14,
    "quinze":15, "dezesseis":16, "dezessete":17, "dezoito":18, "dezenove":19
}
tens = {
    "vinte":20, "trinta":30, "quarenta":40, "cinquenta":50, "sessenta":60,
    "setenta":70, "oitenta":80, "noventa":90
}
hundreds = {
    "cem":100, "cento":100, "duzentos":200, "trezentos":300, "quatrocentos":400,
    "quinhentos":500, "seiscentos":600, "setecentos":700, "oitocentos":800,
    "novecentos":900
}
word2num_all = {**units, **teens, **tens, **hundreds}

def parse_number_words(tokens):
    """Parse list of number-word tokens (no 'e' tokens) into integer (supports up to 999)."""
    total = 0
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in hundreds:
            total += hundreds[t]
            i += 1
        elif t in tens:
            val = tens[t]
            if i + 1 < len(tokens) and tokens[i+1] in units:
                val += units[tokens[i+1]]
                i += 2
            else:
                i += 1
            total += val
        elif t in teens:
            total += teens[t]; i += 1
        elif t in units:
            total += units[t]; i += 1
        else:
            i += 1
    return total if total > 0 else None

def separate_numbers_and_words(text):
    """Insert spaces between digit-word and between number-words glued to words."""
    text = text.lower()
    text = re.sub(r"(\d+)([a-zA-Z])", r"\1 \2", text)
    text = re.sub(r"([a-zA-Z])(\d+)", r"\1 \2", text)

    # Protect compound teen numbers so they don't get split
    protected_teens = ["dezesseis", "dezessete", "dezoito", "dezenove"]
    for teen in protected_teens:
        text = text.replace(teen, f" {teen} ")

    # Now process other number words normally
    keys = sorted(word2num_all.keys(), key=len, reverse=True)
    for w in keys:
        if w not in protected_teens:
            text = re.sub(rf"\b{re.escape(w)}\b", f" {w} ", text)

    text = re.sub(r"\s+", " ", text).strip()
    return text

def extract_numbers_and_positions(tokens):
    """Extract all numbers and their positions from tokens"""
    numbers = []
    
    i = 0
    while i < len(tokens):
        if tokens[i].isdigit():
            numbers.append((i, int(tokens[i])))
            i += 1
        elif tokens[i] in word2num_all:
            # Only combine number words if they're connected by "e"
            num_tokens = [tokens[i]]
            j = i + 1
            
            # Look for "e" followed by a number word
            while j < len(tokens) - 1:
                if tokens[j] == "e" and tokens[j+1] in word2num_all:
                    num_tokens.extend([tokens[j], tokens[j+1]])
                    j += 2
                else:
                    break
            
            # Parse the number tokens
            number = parse_number_words([t for t in num_tokens if t != "e"])
            if number:
                numbers.append((i, number))
                i = j
            else:
                i += 1
        else:
            i += 1
            
    return numbers

def find_associated_number(product_position, all_tokens, numbers_with_positions):
    """Find the number associated with a product based on word order patterns"""
    if not numbers_with_positions:
        return 1, None
    
    # Pattern 1: Number immediately before the product (most common)
    if product_position > 0:
        prev_token = all_tokens[product_position - 1]
        if prev_token.isdigit() or prev_token in word2num_all:
            for pos, val in numbers_with_positions:
                if pos == product_position - 1:
                    return val, pos
    
    # Pattern 2: Look for numbers before the product (anywhere before)
    numbers_before = [(pos, val) for pos, val in numbers_with_positions if pos < product_position]
    if numbers_before:
        # Return the closest number before the product (highest position number before product)
        closest_before = max(numbers_before, key=lambda x: x[0])
        return closest_before[1], closest_before[0]
    
    # Pattern 3: Number immediately after the product
    if product_position + 1 < len(all_tokens):
        next_token = all_tokens[product_position + 1]
        if next_token.isdigit() or next_token in word2num_all:
            for pos, val in numbers_with_positions:
                if pos == product_position + 1:
                    return val, pos
    
    # Pattern 4: Look for numbers after the product (anywhere after)
    numbers_after = [(pos, val) for pos, val in numbers_with_positions if pos > product_position]
    if numbers_after:
        # Return the closest number after the product (lowest position number after product)
        closest_after = min(numbers_after, key=lambda x: x[0])
        return closest_after[1], closest_after[0]
    
    return 1, None

def parse_order_interactive(message, products_db, similarity_threshold=80, uncertain_range=(60, 80)):
    """
    Interactive version that uses pattern-based quantity association with multi-word product support.
    """
    message = normalize(message)
    message = separate_numbers_and_words(message)
    message = re.sub(r"[,\.;\+\-\/\(\)\[\]\:]", " ", message)
    message = re.sub(r"\s+", " ", message).strip()

    tokens = message.split()
    working_db = deepcopy(products_db)
    parsed_orders = []

    # Extract all numbers and their positions
    numbers_with_positions = extract_numbers_and_positions(tokens)
    
    # Sort products by word count (longest first) to prioritize multi-word matches
    sorted_products = sorted(working_db, key=lambda x: len(x[0].split()), reverse=True)
    normalized_products = [normalize(p) for p, _ in sorted_products]
    max_prod_words = max(len(p.split()) for p, _ in sorted_products)

    # Precompute the set of words that appear in any product name
    product_words = set()
    for product, _ in working_db:
        for word in product.split():
            product_words.add(normalize(word))

    i = 0
    while i < len(tokens):
        token = tokens[i]

        # Skip filler words and numbers only if they are not part of a product name
        filler_words = {"quero"}
        if (token in filler_words and token not in product_words) or token.isdigit() or token in word2num_all:
            i += 1
            continue

        matched = False
        confirmed_product = None
        confirmed_quantity = None
        confirmed_size = 0
        confirmed_number_position = None
        
        # Try different phrase lengths (longest first) - prioritize multi-word products
        for size in range(min(max_prod_words, 4), 0, -1):
            if i + size > len(tokens):
                continue
                
            # Skip if any token in the phrase is a number or filler word (unless part of product)
            phrase_tokens = tokens[i:i+size]
            if any((t.isdigit() or t in word2num_all or (t in filler_words and t not in product_words)) for t in phrase_tokens):
                continue
                
            phrase = " ".join(phrase_tokens)
            phrase_norm = normalize(phrase)

            best_score = 0
            best_idx = None
            best_product = None
            
            # Find best match for this phrase length (check against sorted products)
            for idx, prod_norm in enumerate(normalized_products):
                score = similarity_percentage(phrase_norm, prod_norm)
                if score > best_score:
                    best_score = score
                    best_product = sorted_products[idx][0]
                    best_idx = next(j for j, (p, _) in enumerate(working_db) if p == best_product)

            # Handle the match
            if best_score >= similarity_threshold:
                # For web version, we'll auto-confirm matches above threshold
                quantity, number_position = find_associated_number(i, tokens, numbers_with_positions)
                working_db[best_idx][1] += quantity
                parsed_orders.append({"product": best_product, "qty": quantity, "score": round(best_score,2)})
                
                # Remove the used number from available numbers
                if number_position is not None:
                    numbers_with_positions = [(pos, val) for pos, val in numbers_with_positions if pos != number_position]
                
                i += size
                matched = True
                break

        if not matched:
            # If no match found, find the best match to suggest
            phrase = tokens[i]
            best_match = None
            best_score = 0
            phrase_norm = normalize(phrase)
            
            for product, _ in working_db:
                score = similarity_percentage(phrase_norm, normalize(product))
                if score > best_score:
                    best_score = score
                    best_match = product
            
            if best_match and best_score > 50:
                # Auto-confirm reasonable matches for web version
                quantity, number_position = find_associated_number(i, tokens, numbers_with_positions)
                for idx, (product, _) in enumerate(working_db):
                    if product == best_match:
                        working_db[idx][1] += quantity
                        parsed_orders.append({
                            "product": best_match,
                            "qty": quantity,
                            "score": round(best_score, 2)
                        })
                        
                        if number_position is not None:
                            numbers_with_positions = [(pos, val) for pos, val in numbers_with_positions if pos != number_position]
                        
                        matched = True
                        break
            i += 1

    return parsed_orders, working_db

# ---------- OrderBot for Web ----------
class OrderBot:
    def __init__(self, products_db):
        self.products_db = products_db
        self.current_db = deepcopy(products_db)
        self.waiting_for_confirmation = False
        self.order_confirmed = False
        self.order_canceled = False

    def process_text(self, text):
        """Process incoming text. Returns bot response and current orders."""
        text = text.strip().lower()
        
        if self.waiting_for_confirmation:
            return self.process_confirmation(text)
        
        # Check for special commands
        if text in ['pronto', 'finalizar', 'fim']:
            return self.send_order_summary()
        
        if text in ['cancelar', 'hoje não']:
            self.order_canceled = True
            return "OK, volte sempre!", self.get_current_orders()
        
        if text == '/clear':
            self.clear_orders()
            return "Planilha esvaziada com sucesso!", self.get_current_orders()
        
        # Process the order using your proven logic
        parsed, self.current_db = parse_order_interactive(
            text, 
            self.current_db,
            similarity_threshold=80,
            uncertain_range=(60, 80)
        )
        
        response = "Pedido processado:\n"
        for order in parsed:
            response += f"✓ {order['product']}: +{order['qty']} (confiança: {order['score']}%)\n"
        
        if not parsed:
            response = "Nenhum item reconhecido. Por favor, tente novamente com outros termos."
        
        self._save_to_excel()
        return response, self.get_current_orders()

    def process_confirmation(self, response):
        """Process confirmation response"""
        response_lower = response.lower()
        
        if any(word in response_lower for word in ['confirmo', 'sim', 'correto', 'certo', 's']):
            self.order_confirmed = True
            self.waiting_for_confirmation = False
            final_orders = self.get_current_orders()
            self._save_final_orders()
            return "Pedido confirmado com sucesso!", final_orders
        elif any(word in response_lower for word in ['nao', 'não', 'n']):
            self.waiting_for_confirmation = False
            self.current_db = deepcopy(self.products_db)
            return "Por favor, reescreva seu pedido.", self.get_current_orders()
        else:
            return "Resposta não reconhecida. Por favor, responda com 'sim' ou 'nao'.", self.get_current_orders()

    def send_order_summary(self):
        """Send order summary and ask for confirmation"""
        self.waiting_for_confirmation = True
        summary = "\n📋 Resumo do seu pedido:\n"
        orders = self.get_current_orders()
        
        for product, qty in orders.items():
            if qty > 0:
                summary += f"• {product}: {qty}\n"
        
        summary += "\nConfirma o pedido? (responda com 'sim' ou 'nao')"
        return summary, orders

    def get_current_orders(self):
        """Get current orders as dict"""
        return {product: qty for product, qty in self.current_db if qty > 0}

    def clear_orders(self):
        """Clear all orders"""
        self.current_db = deepcopy(self.products_db)
        with excel_lock:
            # Clear the worksheet except header
            ws.delete_rows(2, ws.max_row)
            # Re-initialize with zero quantities
            for product, _ in self.products_db:
                ws.append([product, 0])
            wb.save(EXCEL_PATH)

    def _save_to_excel(self):
        """Save current orders to Excel"""
        with excel_lock:
            # Clear existing data (keep header)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            
            # Add current orders
            for product, qty in self.current_db:
                ws.append([product, qty])
            
            wb.save(EXCEL_PATH)

    def _save_final_orders(self):
        """Save final confirmed orders - this matches your original logic"""
        with excel_lock:
            # Your original saving logic
            for row in range(2, len(self.products_db) + 2):
                product_name = self.current_db[row-2][0]
                current_qty = self.current_db[row-2][1]
                
                # Find or create row for this product
                product_found = False
                for excel_row in range(2, ws.max_row + 1):
                    if ws[f'A{excel_row}'].value == product_name:
                        # Update existing quantity
                        existing_qty = ws[f'B{excel_row}'].value or 0
                        ws[f'B{excel_row}'] = existing_qty + current_qty
                        product_found = True
                        break
                
                if not product_found:
                    # Add new row
                    ws.append([product_name, current_qty])
            
            wb.save(EXCEL_PATH)

# ---------- Initialize products_db ----------
products_db = [
    ["abacaxi", 0], ["abacaxi com hortela", 0], ["acai", 0], ["acerola", 0],
    ["ameixa", 0], ["caja", 0], ["caju", 0], ["goiaba", 0], ["graviola", 0],
    ["manga", 0], ["maracuja", 0], ["morango", 0], ["seriguela", 0], ["tamarindo", 0],
    ["caixa de ovos", 0], ["ovo", 0], ["queijo", 0]
]

# Store bots by session
bots = {}

def get_bot(session_id):
    """Get or create bot for session"""
    if session_id not in bots:
        bots[session_id] = OrderBot(products_db)
    return bots[session_id]

# ---------- Flask routes ----------
@app.route("/")
def index():
    session_id = request.args.get('session_id', 'default')
    return render_template("index.html", session_id=session_id)

@app.route("/send_message", methods=["POST"])
def send_message():
    data = request.json
    text = data.get("message", "")
    session_id = data.get("session_id", "default")
    
    bot = get_bot(session_id)
    response, orders = bot.process_text(text)
    
    return jsonify({"response": response, "orders": orders})

@app.route("/get_orders", methods=["GET"])
def get_orders():
    session_id = request.args.get("session_id", "default")
    bot = get_bot(session_id)
    orders = bot.get_current_orders()
    return jsonify(orders)

@app.route("/download_excel", methods=["GET"])
def download_excel():
    return send_file(EXCEL_PATH, as_attachment=True, download_name='pedidos.xlsx')

if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000)